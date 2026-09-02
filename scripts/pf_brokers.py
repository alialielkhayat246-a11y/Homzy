"""Scrape PropertyFinder Egypt's company/broker directory into an Excel sheet.

Kept LOCAL (data/*.xlsx is gitignored) — it contains thousands of phone numbers.
Source: the public find-broker directory + each company's page __NEXT_DATA__.
"""
import json, re, sys, time, urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120 Safari/537.36")
BASE = "https://www.propertyfinder.eg"
ND = re.compile(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', re.S)


def _get(url, want_json=False, tries=3):
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": UA, "Accept-Language": "en",
                "Accept": "application/json" if want_json else "text/html"})
            raw = urllib.request.urlopen(req, timeout=25).read().decode("utf-8", "ignore")
            return json.loads(raw) if want_json else raw
        except Exception:
            time.sleep(1.2 * (i + 1))
    return None


def company_list():
    d = _get(BASE + "/api/pwa/brokerages/list?locale=en", want_json=True)
    return [{"id": str(x["i"]), "name": x["n"]} for x in (d or [])]


def _agents_array(pp):
    ag = pp.get("agents")
    if isinstance(ag, list):
        return ag
    if isinstance(ag, dict):
        if isinstance(ag.get("data"), list):
            return ag["data"]
        for v in ag.values():
            if isinstance(v, list):
                return v
    return []


def scrape_one(co):
    html = _get(f"{BASE}/en/broker/{co['id']}")
    if not html:
        return None
    m = ND.search(html)
    if not m:
        return None
    try:
        pp = json.loads(m.group(1))["props"]["pageProps"]
    except Exception:
        return None
    b = pp.get("broker") or {}
    link = f"{BASE}/en/broker/{co['id']}"
    company = {
        "name": b.get("name") or co["name"],
        "phone": b.get("phone") or "",
        "units": b.get("totalProperties") or 0,
        "sale": (b.get("propertiesResidentialForSaleCount") or 0) + (b.get("propertiesCommercialForSaleCount") or 0),
        "rent": (b.get("propertiesResidentialForRentCount") or 0) + (b.get("propertiesCommercialForRentCount") or 0),
        "agents": b.get("totalAgents") or 0,
        "address": (b.get("address") or "").strip(),
        "link": link,
    }
    agents = []
    for a in _agents_array(pp):
        slug = a.get("slug") or ""
        agents.append({
            "company": company["name"],
            "name": a.get("name") or "",
            "units": a.get("totalProperties") or 0,
            "phone": a.get("phone") or a.get("whatsappPhone") or "",
            "whatsapp": a.get("whatsappPhone") or "",
            "link": f"{BASE}/en/broker/{co['id']}/{slug}" if slug else link,
        })
    return company, agents


def main():
    cos = company_list()
    print(f"companies to fetch: {len(cos)}", flush=True)
    companies, agents, fails = [], [], 0
    done = 0
    with ThreadPoolExecutor(max_workers=12) as ex:
        futs = {ex.submit(scrape_one, c): c for c in cos}
        for f in as_completed(futs):
            done += 1
            r = f.result()
            if r is None:
                fails += 1
            else:
                companies.append(r[0]); agents.extend(r[1])
            if done % 100 == 0:
                print(f"  {done}/{len(cos)} (agents={len(agents)}, fails={fails})", flush=True)
    print(f"done: {len(companies)} companies, {len(agents)} agents, {fails} fails", flush=True)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="0B5563")
    hdr_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncols):
        ws.sheet_view.rightToLeft = True
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    # Sheet 1 — Marketers (the requested columns, agent-level)
    ws1 = wb.active; ws1.title = "المسوّقين"
    ws1.append(["اسم الشركة", "اسم المسوّق", "عدد الوحدات", "رقم الموبايل", "لينك الصفحة"])
    agents.sort(key=lambda a: (a["company"], -int(a["units"] or 0)))
    for a in agents:
        ws1.append([a["company"], a["name"], a["units"], a["phone"], a["link"]])
    style_header(ws1, 5)
    for col, w in zip("ABCDE", [30, 26, 12, 18, 60]):
        ws1.column_dimensions[col].width = w

    # Sheet 2 — Companies summary
    ws2 = wb.create_sheet("الشركات")
    ws2.append(["اسم الشركة", "رقم موبايل الشركة", "إجمالي الوحدات", "للبيع",
                "للإيجار", "عدد المسوّقين", "العنوان", "لينك الصفحة"])
    companies.sort(key=lambda c: -int(c["units"] or 0))
    for c in companies:
        ws2.append([c["name"], c["phone"], c["units"], c["sale"], c["rent"],
                    c["agents"], c["address"], c["link"]])
    style_header(ws2, 8)
    for col, w in zip("ABCDEFGH", [30, 18, 14, 10, 10, 12, 50, 55]):
        ws2.column_dimensions[col].width = w

    out = "data/propertyfinder_brokers.xlsx"
    wb.save(out)
    print(f"saved: {out}  |  {len(companies)} companies, {len(agents)} marketers", flush=True)


if __name__ == "__main__":
    main()
