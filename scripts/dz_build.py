"""Build the Dubizzle verified-agencies Excel from the scraped JSON.
Kept LOCAL (data/*.xlsx is gitignored)."""
import json, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = sys.argv[1] if len(sys.argv) > 1 else None
with open(SRC, encoding="utf-8") as fh:
    rows = json.load(fh)


def clean_phone(p):
    if not p:
        return "غير متاح علنًا"
    d = re.sub(r"[^\d]", "", p)
    if not d or re.search(r"0{6,}", d) or len(d) < 8:
        return "غير متاح علنًا"
    return p


def paid_label(r):
    if r["pr"] in ("elite", "featured", "premium", "hero") or r.get("vip"):
        return "نعم (" + (r["pr"] or "VIP") + ")"
    if r["ap"] >= 20:
        return "غالبًا (نشاط عالٍ)"
    return "-"


rows.sort(key=lambda r: -int(r.get("a") or 0))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "شركات موثّقة"
ws.sheet_view.rightToLeft = True

headers = ["#", "اسم الشركة", "موثّق", "إعلانات مدفوعة؟", "رقم الموبايل",
           "إجمالي الإعلانات", "للبيع", "للإيجار", "رقم الرخصة",
           "نشاط (ظهور)", "لينك الصفحة"]
ws.append(headers)

hdr_fill = PatternFill("solid", fgColor="0B5563")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
alt = PatternFill("solid", fgColor="E2EEF0")
link_font = Font(color="0B5563", underline="single")
thin = Side(style="thin", color="D9D2CB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, r in enumerate(rows, 1):
    ws.append([
        i, r["n"], "نعم", paid_label(r), clean_phone(r.get("p")),
        int(r.get("a") or 0), int(r.get("s") or 0), int(r.get("e") or 0),
        r.get("lic") or "", int(r.get("ap") or 0), r["l"],
    ])

for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 30

for ri in range(2, ws.max_row + 1):
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.border = border
        cell.alignment = Alignment(vertical="center",
                                   horizontal="left" if ci == 11 else "center")
        if ci == 11:
            cell.font = link_font
        if ri % 2 == 0:
            cell.fill = alt

for col, w in zip("ABCDEFGHIJK", [5, 34, 8, 20, 18, 15, 9, 9, 16, 12, 60]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

out = "data/dubizzle_verified_agencies.xlsx"
wb.save(out)
withph = sum(1 for r in rows if clean_phone(r.get("p")) != "غير متاح علنًا")
print(f"saved: {out} | {len(rows)} verified agencies | {withph} with public phone")
