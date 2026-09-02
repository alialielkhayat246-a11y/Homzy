"""Fetch ALL marketers (agents) for every PropertyFinder Egypt company via the
public agents API (paginated), and rebuild the Excel. Kept LOCAL (gitignored).

Sheet 1 "المسوّقين": every agent — company, agent name, units, mobile, link.
Sheet 2 "الشركات": company summary (reused from the previous run's workbook).
"""
import json, urllib.request, urllib.parse, time
from concurrent.futures import ThreadPoolExecutor, as_completed

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120 Safari/537.36")
BASE = "https://www.propertyfinder.eg"


def get_json(url, tries=3):
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": UA, "Accept": "application/json", "Accept-Language": "en"})
            return json.loads(urllib.request.urlopen(req, timeout=25).read().decode("utf-8", "ignore"))
        except Exception:
            time.sleep(1.0 * (i + 1))
    return None


def company_list():
    d = get_json(BASE + "/api/pwa/brokerages/list?locale=en") or []
    return [(str(x["i"]), x["n"]) for x in d]


def agents_page(cid, page):
    return get_json(f"{BASE}/api/pwa/agents?broker_id={cid}&page={page}")


def collect(cid, name):
    """All agents for one company (follow pagination)."""
    out = []
    j = agents_page(cid, 1)
    if not j:
        return out, True  # failed
    pages = (j.get("meta") or {}).get("totalPages", 1) or 1
    batches = [j.get("agents", [])]
    for p in range(2, pages + 1):
        jp = agents_page(cid, p)
        batches.append((jp or {}).get("agents", []))
    for arr in batches:
        for a in arr:
            slug = a.get("slug") or ""
            out.append([
                name,
                a.get("name") or "",
                a.get("totalProperties") or 0,
                a.get("phone") or a.get("whatsappPhone") or "",
                f"{BASE}/en/broker/{cid}/{slug}" if slug else f"{BASE}/en/broker/{cid}",
            ])
    return out, False


def main():
    cos = company_list()
    print(f"companies: {len(cos)}", flush=True)
    rows, fails, done = [], 0, 0
    with ThreadPoolExecutor(max_workers=12) as ex:
        futs = {ex.submit(collect, cid, name): cid for cid, name in cos}
        for f in as_completed(futs):
            done += 1
            ags, failed = f.result()
            if failed:
                fails += 1
            rows.extend(ags)
            if done % 100 == 0:
                print(f"  {done}/{len(cos)} agents={len(rows)} fails={fails}", flush=True)
    print(f"done: {len(rows)} agents, {fails} company-fails", flush=True)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    src = "data/propertyfinder_brokers.xlsx"
    wb = openpyxl.load_workbook(src)  # keep the الشركات sheet as-is
    if "المسوّقين" in wb.sheetnames:
        del wb["المسوّقين"]
    ws = wb.create_sheet("المسوّقين", 0)
    ws.sheet_view.rightToLeft = True
    ws.append(["اسم الشركة", "اسم المسوّق", "عدد الوحدات", "رقم الموبايل", "لينك الصفحة"])
    # sort: company, then most units first
    rows.sort(key=lambda r: (str(r[0]), -int(r[2] or 0)))
    for r in rows:
        ws.append(r)
    fill = PatternFill("solid", fgColor="0B5563"); font = Font(bold=True, color="FFFFFF")
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c); cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDE", [30, 26, 12, 18, 60]):
        ws.column_dimensions[col].width = w

    wb.save(src)
    print(f"saved: {src}  |  {len(rows)} marketers across {len(cos)} companies", flush=True)


if __name__ == "__main__":
    main()
